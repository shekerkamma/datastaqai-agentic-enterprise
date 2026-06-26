#!/usr/bin/env python3
"""Build the Excel competitor database from the vendor + use-case datasets."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

A=Path("/mnt/d/New folder/Antigravity-test/antigravity-skills/wedge_dossiers/analysis")
H=Path("/mnt/d/New folder/Antigravity-test/antigravity-skills/wedge_dossiers/book")
NAVY="0A1628"; TEAL="00C9A7"; SOFT="F4F7F8"
hdr_fill=PatternFill("solid",fgColor=NAVY); hdr_font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
thin=Side(style="thin",color="D9DFE5"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def sheet(wb,name,rows,title):
    ws=wb.create_sheet(name)
    ws.append([title]); ws["A1"].font=Font(bold=True,size=14,color=NAVY)
    ws.append([])
    head=rows[0]
    ws.append(head)
    hr=ws.max_row
    for j,_ in enumerate(head,1):
        c=ws.cell(hr,j); c.fill=hdr_fill; c.font=hdr_font; c.alignment=Alignment(wrap_text=True,vertical="center"); c.border=border
    for r in rows[1:]:
        ws.append(r)
        for j in range(1,len(head)+1):
            cc=ws.cell(ws.max_row,j); cc.border=border; cc.alignment=Alignment(wrap_text=True,vertical="top")
            if ws.max_row%2==0: cc.fill=PatternFill("solid",fgColor=SOFT)
    ws.freeze_panes=ws.cell(hr+1,1)
    for j,_ in enumerate(head,1):
        w=max(12,min(46,max((len(str(rows[k][j-1])) for k in range(len(rows)) if j-1<len(rows[k])),default=12)+2))
        ws.column_dimensions[get_column_letter(j)].width=w
    return ws

def read(path):
    with open(path,encoding="utf-8") as f:
        return [list(r) for r in csv.reader(f)]

wb=Workbook(); wb.remove(wb.active)
# Vendors
vend=read(A/"vendors.csv")
sheet(wb,"Vendors",vend,"DataStaqAI — Competitor Database · Vendors (source-traceable; blanks = not publicly disclosed)")
# Use cases
uc=read(A/"competitive_dataset.csv")
sheet(wb,"Use Cases",uc,"Agentic Use Cases · posture, pattern, TAM, price floor, competitive density (density = judgment)")
# Funding-by-cluster summary
from collections import defaultdict
hdr=vend[0]; idx={h:i for i,h in enumerate(hdr)}
raised=defaultdict(float); val=defaultdict(list); n=defaultdict(int)
for r in vend[1:]:
    cl=r[idx["cluster"]]; n[cl]+=1
    try: raised[cl]+=float(r[idx["total_raised_usd_m"]])
    except: pass
    try: val[cl].append(float(r[idx["valuation_usd_b"]]))
    except: pass
summ=[["Cluster","Vendors","Disclosed raised ($M)","Top valuation ($B)"]]
for cl in sorted(raised,key=lambda k:-raised[k]):
    summ.append([cl,n[cl],round(raised[cl]),max(val[cl]) if val[cl] else ""])
sheet(wb,"Funding Summary",summ,"Disclosed funding & valuation by cluster")
# README
ws=wb.create_sheet("README",0)
ws["A1"]="DataStaqAI — Competitor Database"; ws["A1"].font=Font(bold=True,size=16,color=NAVY)
notes=["Companion to 'The Agentic Enterprise' (48-slide deck + Word report).",
       "Sheets: Vendors (42, named) · Use Cases (25) · Funding Summary (by cluster).",
       "Every figure traces to the deep-research files (COMPETITIVE_*.md). Blanks = not publicly disclosed (not zero).",
       "Competitive density and DataStaqAI posture are labeled judgment, not measured indices.",
       "Generated 2026 · figures current to the research run."]
for i,t in enumerate(notes,3): ws.cell(i,1,t)
ws.column_dimensions["A"].width=110

out=H/"DataStaqAI-Competitor-Database.xlsx"
wb.save(str(out)); print("EXCEL:",out,"| sheets:",wb.sheetnames)
