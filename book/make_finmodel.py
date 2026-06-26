#!/usr/bin/env python3
"""#8 — bottoms-up 3-year financial model (.xlsx, live formulas + editable assumptions)."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
H=Path("/mnt/d/New folder/Antigravity-test/antigravity-skills/wedge_dossiers/book")
NAVY="0A1628"; TEAL="00C9A7"; GOLD="FFB800"; SOFT="F4F7F8"; INPUT="FFF6D9"
hf=PatternFill("solid",fgColor=NAVY); hfont=Font(bold=True,color="FFFFFF")
inpfill=PatternFill("solid",fgColor=INPUT); thin=Side(style="thin",color="D9DFE5"); bd=Border(thin,thin,thin,thin)
wb=Workbook()

# ── Assumptions (editable inputs highlighted) ──
A=wb.active; A.title="Assumptions"
A["A1"]="DataStaqAI — Financial Model · Assumptions (edit yellow cells; Model recalculates)"; A["A1"].font=Font(bold=True,size=13,color=NAVY)
rows=[("NRR_annual","Net revenue retention (annual)",1.35,"x"),
 ("Landing_ACV","Avg landing ACV (recurring, $/yr)",40000,"$"),
 ("GM_Y1","Gross margin — Year 1",0.52,"%"),("GM_Y2","Gross margin — Year 2",0.62,"%"),("GM_Y3","Gross margin — Year 3",0.70,"%"),
 ("Logos_Y1","New logos — Year 1",15,"#"),("Logos_Y2","New logos — Year 2",35,"#"),("Logos_Y3","New logos — Year 3",45,"#"),
 ("Build_fee","Avg BUILD project fee ($)",80000,"$"),
 ("Build_Y1","BUILD projects — Year 1",5,"#"),("Build_Y2","BUILD projects — Year 2",10,"#"),("Build_Y3","BUILD projects — Year 3",14,"#"),
 ("Accts_per_pod","Accounts per delivery pod",8,"#"),("People_per_pod","People per pod",2,"#"),
 ("Loaded_cost","Fully-loaded cost / person ($)",160000,"$"),
 ("Overhead_Y1","Non-pod headcount — Y1",3,"#"),("Overhead_Y2","Non-pod headcount — Y2",4,"#"),("Overhead_Y3","Non-pod headcount — Y3",5,"#"),
 ("CAC","Cash CAC per new logo ($)",8000,"$"),("Opex_pct","Opex (S&M+G&A) as % of revenue",0.30,"%"),
 ("Avg_life_yrs","Avg customer life (yrs, for LTV)",4,"#")]
A.append([]); A.append(["Key","Driver","Value","Unit"])
for j in range(1,5): c=A.cell(3,j); c.fill=hf; c.font=hfont; c.border=bd
namecell={}
for i,(k,lab,v,u) in enumerate(rows,start=4):
    A.cell(i,1,k); A.cell(i,2,lab); vc=A.cell(i,3,v); A.cell(i,4,u)
    vc.fill=inpfill; vc.border=bd
    if u=="%": vc.number_format="0%"
    elif u=="$": vc.number_format="#,##0"
    elif u=="x": vc.number_format="0.00"
    namecell[k]=f"Assumptions!$C${i}"
A.column_dimensions["A"].width=16; A.column_dimensions["B"].width=40; A.column_dimensions["C"].width=14; A.column_dimensions["D"].width=8

# ── Model (Y1/Y2/Y3 with live formulas) ──
M=wb.create_sheet("Model")
M["A1"]="3-Year P&L & SaaS metrics (driver-based; edit Assumptions to recalc)"; M["A1"].font=Font(bold=True,size=13,color=NAVY)
M.append([]); M.append(["Line","Year 1","Year 2","Year 3"])
for j in range(1,5): c=M.cell(3,j); c.fill=hf; c.font=hfont; c.border=bd
N=namecell
def gm(col): return {"B":N["GM_Y1"],"C":N["GM_Y2"],"D":N["GM_Y3"]}[col]
def logos(col): return {"B":N["Logos_Y1"],"C":N["Logos_Y2"],"D":N["Logos_Y3"]}[col]
def builds(col): return {"B":N["Build_Y1"],"C":N["Build_Y2"],"D":N["Build_Y3"]}[col]
def oh(col): return {"B":N["Overhead_Y1"],"C":N["Overhead_Y2"],"D":N["Overhead_Y3"]}[col]
# row builder: label, formula(col)->str ; cols B,C,D
def addrow(label,fB,fC,fD,fmt="#,##0",pct=False):
    r=M.max_row+1; M.cell(r,1,label)
    for col,f in zip("BCD",(fB,fC,fD)):
        cell=M[f"{col}{r}"]; cell.value=f; cell.number_format=("0%" if pct else fmt); cell.border=bd
    M.cell(r,1).border=bd
    return r
rN={}
rN["new_logos"]=addrow("New logos",f"={logos('B')}",f"={logos('C')}",f"={logos('D')}",fmt="0")
rN["cum_logos"]=addrow("Cumulative logos",f"=B{rN['new_logos']}",f"=B{rN['cum_logos'] if False else rN['new_logos']}+C{rN['new_logos']}",f"=C{rN['new_logos']}+B{rN['new_logos']}+D{rN['new_logos']}",fmt="0")
# fix cumulative properly
M[f"B{rN['cum_logos']}"]=f"=B{rN['new_logos']}"
M[f"C{rN['cum_logos']}"]=f"=B{rN['cum_logos']}+C{rN['new_logos']}"
M[f"D{rN['cum_logos']}"]=f"=C{rN['cum_logos']}+D{rN['new_logos']}"
rN["new_arr"]=addrow("New recurring ARR ($)",f"=B{rN['new_logos']}*{N['Landing_ACV']}",f"=C{rN['new_logos']}*{N['Landing_ACV']}",f"=D{rN['new_logos']}*{N['Landing_ACV']}")
rN["beg_arr"]=addrow("Beginning recurring ARR ($)","=0",None,None)
rN["exp_arr"]=addrow("Expansion ARR ($)",None,None,None)
rN["end_arr"]=addrow("Ending recurring ARR ($)",None,None,None)
# beginning = prior ending
M[f"C{rN['beg_arr']}"]=f"=B{rN['end_arr']}"; M[f"D{rN['beg_arr']}"]=f"=C{rN['end_arr']}"
for col in "BCD":
    M[f"{col}{rN['exp_arr']}"]=f"={col}{rN['beg_arr']}*({N['NRR_annual']}-1)"
    M[f"{col}{rN['end_arr']}"]=f"={col}{rN['beg_arr']}+{col}{rN['new_arr']}+{col}{rN['exp_arr']}"
rN["rec_rev"]=addrow("Recurring revenue (recognized, $)",None,None,None)
for col in "BCD": M[f"{col}{rN['rec_rev']}"]=f"=({col}{rN['beg_arr']}+{col}{rN['end_arr']})/2"
rN["build_rev"]=addrow("BUILD revenue ($)",f"={builds('B')}*{N['Build_fee']}",f"={builds('C')}*{N['Build_fee']}",f"={builds('D')}*{N['Build_fee']}")
rN["tot_rev"]=addrow("Total revenue ($)",None,None,None)
for col in "BCD": M[f"{col}{rN['tot_rev']}"]=f"={col}{rN['rec_rev']}+{col}{rN['build_rev']}"
rN["gm"]=addrow("Gross margin %",f"={gm('B')}",f"={gm('C')}",f"={gm('D')}",pct=True)
rN["gp"]=addrow("Gross profit ($)",None,None,None)
for col in "BCD": M[f"{col}{rN['gp']}"]=f"={col}{rN['tot_rev']}*{col}{rN['gm']}"
rN["opex"]=addrow("Opex S&M+G&A ($)",None,None,None)
for col in "BCD": M[f"{col}{rN['opex']}"]=f"={col}{rN['tot_rev']}*{N['Opex_pct']}"
rN["ebitda"]=addrow("EBITDA ($)",None,None,None)
for col in "BCD": M[f"{col}{rN['ebitda']}"]=f"={col}{rN['gp']}-{col}{rN['opex']}"
# headcount + metrics
rN["pods"]=addrow("Delivery pods",f"=ROUNDUP(B{rN['cum_logos']}/{N['Accts_per_pod']},0)",f"=ROUNDUP(C{rN['cum_logos']}/{N['Accts_per_pod']},0)",f"=ROUNDUP(D{rN['cum_logos']}/{N['Accts_per_pod']},0)",fmt="0")
rN["head"]=addrow("Total headcount",f"=B{rN['pods']}*{N['People_per_pod']}+{oh('B')}",f"=C{rN['pods']}*{N['People_per_pod']}+{oh('C')}",f"=D{rN['pods']}*{N['People_per_pod']}+{oh('D')}",fmt="0")
rN["rph"]=addrow("Revenue / head ($)",None,None,None)
for col in "BCD": M[f"{col}{rN['rph']}"]=f"={col}{rN['tot_rev']}/{col}{rN['head']}"
rN["cacpb"]=addrow("CAC payback (months)",None,None,None)
for col in "BCD": M[f"{col}{rN['cacpb']}"]=f"={N['CAC']}/(({N['Landing_ACV']}*{col}{rN['gm']})/12)"
M[f"B{rN['cacpb']}"].number_format="0.0"; M[f"C{rN['cacpb']}"].number_format="0.0"; M[f"D{rN['cacpb']}"].number_format="0.0"
rN["ltvcac"]=addrow("LTV / CAC (x)",None,None,None)
for col in "BCD": M[f"{col}{rN['ltvcac']}"]=f"=({N['Landing_ACV']}*{col}{rN['gm']}*{N['Avg_life_yrs']})/{N['CAC']}"
M[f"B{rN['ltvcac']}"].number_format="0.0"; M[f"C{rN['ltvcac']}"].number_format="0.0"; M[f"D{rN['ltvcac']}"].number_format="0.0"
# emphasize key rows
for rr in (rN["end_arr"],rN["tot_rev"],rN["ebitda"]):
    for col in "ABCD":
        M[f"{col}{rr}"].font=Font(bold=True)
        M[f"{col}{rr}"].fill=PatternFill("solid",fgColor=SOFT)
for col in "ABCD": M.column_dimensions[col].width=30 if col=="A" else 15

# ── Notes ──
NT=wb.create_sheet("Notes",0)
NT["A1"]="DataStaqAI Financial Model — read me"; NT["A1"].font=Font(bold=True,size=15,color=NAVY)
notes=["ILLUSTRATIVE planning case. Edit the yellow cells on 'Assumptions' and the 'Model' sheet recalculates live.",
 "Default case → Ending recurring ARR ~ $4.8M and total revenue ~ $4.6M by Year 3 (exit run-rate in the $5–10M band).",
 "$5–10M is sensitive to THREE levers: new logos x landing ACV x NRR. Two routes get there:",
 "   (a) volume route: ~90–95 logos at $40k landing + 35% NRR (the default), or",
 "   (b) expansion route: ~50 logos at $40k landing + ~50% NRR (fewer logos, higher net retention).",
 "Gross margin ramps 52%→70% as the reusable agent/connector library matures (the services-to-software trajectory).",
 "COGS is captured inside gross margin %; headcount is shown for revenue/head sanity, not double-counted in EBITDA.",
 "All figures are a model, not a forecast — replace with real win/loss, cost-to-serve and cohort data before external use.",
 "Companion to 'The Agentic Enterprise' deck + report + competitor database."]
for i,t in enumerate(notes,3): NT.cell(i,1,t)
NT.column_dimensions["A"].width=120
out=H/"DataStaqAI-Financial-Model.xlsx"; wb.save(str(out)); print("MODEL:",out,"| sheets:",wb.sheetnames)
